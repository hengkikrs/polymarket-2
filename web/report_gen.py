"""Report generation for the END_WINDOW bot."""
from __future__ import annotations

import time
from datetime import datetime
from io import BytesIO

import core.state as st
from strategies import enabled_strategies


DEFAULT_HOURS = 24


def _period_label(hours: int) -> str:
    return "1 hari terakhir" if hours == 24 else f"{hours} jam terakhir"


def get_report_data(hours: int = DEFAULT_HOURS) -> dict:
    try:
        hours = int(hours)
    except (TypeError, ValueError):
        hours = DEFAULT_HOURS
    hours = max(1, min(168, hours))
    cutoff = time.time() - hours * 3600
    trades = [t for t in st.load_trades() if float(t.get("timestamp", 0) or 0) >= cutoff]
    resolved = [t for t in trades if t.get("resolved")]
    wins = [t for t in resolved if t.get("won") is True]
    losses = [t for t in resolved if t.get("won") is False]
    total_pnl = round(sum(float(t.get("pnl") or 0.0) for t in resolved), 4)
    total_wagered = round(sum(float(t.get("amount_usd") or 0.0) for t in trades), 4)
    balance = st.load_balance()

    trigs: dict[str, dict] = {}
    for key in ("END_WINDOW",):
        group = [t for t in trades if str(t.get("trigger") or "").upper() == key]
        group_resolved = [t for t in group if t.get("resolved")]
        group_wins = [t for t in group_resolved if t.get("won") is True]
        pnl = round(sum(float(t.get("pnl") or 0.0) for t in group_resolved), 4)
        trigs[key] = {
            "count": len(group),
            "wins": len(group_wins),
            "losses": len(group_resolved) - len(group_wins),
            "pnl": pnl,
            "wr": (len(group_wins) / len(group_resolved) * 100) if group_resolved else 0.0,
            "avg_entry": round(sum(float(t.get("entry_price") or 0.0) for t in group) / len(group), 4) if group else 0.0,
            "avg_delta": round(sum(float(t.get("btc_distance") or 0.0) for t in group) / len(group), 1) if group else 0.0,
        }

    return {
        "hours": hours,
        "period": _period_label(hours),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "strategy_mode": "+".join(sorted(enabled_strategies())) or "END_WINDOW",
        "trades": trades,
        "total": len(trades),
        "resolved": len(resolved),
        "wins": len(wins),
        "losses": len(losses),
        "win_rate": (len(wins) / len(resolved) * 100) if resolved else 0.0,
        "total_pnl": total_pnl,
        "total_wagered": total_wagered,
        "balance": round(float(balance.get("balance", 0.0) or 0.0), 2),
        "initial": round(float(balance.get("initial", 0.0) or 0.0), 2),
        "return_pct": round(
            (float(balance.get("balance", 0.0) or 0.0) - float(balance.get("initial", 1.0) or 1.0))
            / max(float(balance.get("initial", 1.0) or 1.0), 1.0)
            * 100,
            2,
        ),
        "trigs": trigs,
    }


def build_narrative(data: dict) -> list[str]:
    if data["total"] == 0:
        return [f"Tidak ada trade tercatat dalam {data['period']}."]
    verdict = "POSITIF" if data["total_pnl"] >= 0 else "NEGATIF"
    return [
        (
            f"Dalam {data['period']}, bot mencatat {data['total']} trade "
            f"({data['resolved']} resolved) dengan hasil {verdict}: "
            f"P&L ${data['total_pnl']:+.2f}, win rate {data['win_rate']:.1f}%."
        )
    ]


def build_advice(data: dict) -> list[str]:
    advice: list[str] = []
    if data["total"] == 0:
        advice.append("Pastikan dashboard START aktif dan market BTC 5m tersedia.")
    if data["balance"] <= 0:
        advice.append("Saldo mock habis; reset mock atau tunggu trade terbuka selesai resolve.")
    if data["resolved"] >= 5 and data["win_rate"] < 50:
        advice.append("Win rate rendah; evaluasi ulang threshold delta dan batas harga END_WINDOW.")
    if not advice:
        advice.append("Pantau hasil per minimal 20 trade sebelum menarik kesimpulan.")
    return advice


def generate_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Generated", data["generated_at"]),
        ("Mode", data["strategy_mode"]),
        ("Period", data["period"]),
        ("Trades", data["total"]),
        ("Resolved", data["resolved"]),
        ("Wins", data["wins"]),
        ("Losses", data["losses"]),
        ("Win Rate", f"{data['win_rate']:.1f}%"),
        ("P&L", data["total_pnl"]),
        ("Balance", data["balance"]),
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True)

    wt = wb.create_sheet("Trades")
    wt.append(["Time", "Trigger", "Outcome", "Entry", "BTC Delta", "Resolved", "Won", "P&L", "Reason"])
    for cell in wt[1]:
        cell.font = Font(bold=True)
    for trade in data["trades"]:
        ts = datetime.fromtimestamp(float(trade.get("timestamp", 0) or 0)).strftime("%Y-%m-%d %H:%M:%S")
        wt.append([
            ts,
            trade.get("trigger", ""),
            trade.get("outcome", ""),
            float(trade.get("entry_price", 0.0) or 0.0),
            float(trade.get("btc_distance", 0.0) or 0.0),
            bool(trade.get("resolved")),
            trade.get("won"),
            float(trade.get("pnl", 0.0) or 0.0),
            trade.get("trigger_reason", ""),
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("END_WINDOW Bot Report", styles["Title"]),
        Paragraph(f"{data['period']} | generated {data['generated_at']}", styles["Normal"]),
        Spacer(1, 12),
    ]
    summary = [
        ["Mode", data["strategy_mode"]],
        ["Trades", str(data["total"])],
        ["Resolved", str(data["resolved"])],
        ["Win Rate", f"{data['win_rate']:.1f}%"],
        ["P&L", f"${data['total_pnl']:+.2f}"],
        ["Balance", f"${data['balance']:.2f}"],
    ]
    story.append(Table(summary))
    story.append(Spacer(1, 12))
    for paragraph in build_narrative(data):
        story.append(Paragraph(paragraph, styles["BodyText"]))
    story.append(Spacer(1, 8))
    for item in build_advice(data):
        story.append(Paragraph(f"- {item}", styles["BodyText"]))
    doc.build(story)
    return buf.getvalue()
